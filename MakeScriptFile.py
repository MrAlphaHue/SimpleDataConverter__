import time
import openpyxl
import PathManage
import glob
import base64
import os;

conststr_comment = """
//Warn : Don't change this code.
//Generated By MrHue.SimpleDataConverter

"""

conststr_namespace = """
using System;
using System.Collections;
using System.Collections.Generic;
using UnityEngine;
using MrHue;

"""

conststr_initInt = "public int {name};\n"
conststr_initLong = "public long {name};\n"
conststr_initFloat = "public float {name};\n"
conststr_initDouble = "public double {name};\n"
conststr_initString = "public string {name};\n"
conststr_initBool = "public bool {name};\n"
conststr_initXInt = "public XInt {name};\n"

conststr_startBracket = "{"
conststr_endBracket   = "}"
conststr_tableClass = """
public static class {table_name}
{{
    {content}
}}
"""

#cv = convert
conststr_data_cvtoInt = "data.{0} = int.Parse(data_string[{1}]);\n"
conststr_data_cvtoLong = "data.{0} = long.Parse(data_string[{1}]);\n"
conststr_data_cvtoFloat = "data.{0} = float.Parse(data_string[{1}]);\n"
conststr_data_cvtoDouble = "data.{0} = Double.Parse(data_string[{1}]);\n"
conststr_data_cvtoString = """data.{0} = data_string[{1}].Replace("\\\\n", "\\n");\n"""
conststr_data_cvtoBool = """data.{0} = data_string[{1}] == "1" ? true : false;\n"""
conststr_data_cvtoXint = "data.{0} = new XInt(data_string[{1}]);\n"

#debug
conststr_debugInt = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugLong = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugFloat = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugDouble = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugString = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugBool = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString());\n"""
conststr_debugXint = """str += string.Format("{{0}} : {{1}}\\n", "{name}", {name}.ToString_1eN());\n"""


conststr_sheetRegion = """
    #region SheetName_Struct
    public class {sheet_name}_Struct
    {{
        public int struct_index;
{data_struct}

    public string GetAllVars(){{
        string str = "";
{data_debug}
        return str;
        }}
    }}
    //int, long, double, float, 
    public class {sheet_name}_Holder : TableSheet_Holder
    {{
        private Dictionary<int, {sheet_name}_Struct> _dic = new Dictionary<int, {sheet_name}_Struct>();
        public readonly Dictionary<int, {sheet_name}_Struct> pDic = new Dictionary<int, {sheet_name}_Struct>();
        
        public int Count
        {{
            get 
            {{
                return _dic.Count; 
            }}
        }}

        public void DebugAllEls()
        {{
            Debug.LogError("{sheet_name} : DebugAll Els");
            foreach (var pair in _dic)
            {{
                Debug.LogError(pair.Key + " : " + pair.Value.GetAllVars());
            }}
        }}

        public {sheet_name}_Struct GetStruct(int index)
        {{
            if (_dic.ContainsKey(index))
                return _dic[index];
            return null;
        }}
        public {sheet_name}_Holder(string[][] strs) : base(strs)
        {{
            this.SheetName = GetType().Name;
            for (int i = 0; i < strs.Length; i++)
            {{
                string[] data_string = strs[i];
                var data = new {sheet_name}_Struct() 
                {{     
                
                }};
                data.struct_index = i;
{data_init}

                _dic.Add(data.struct_index , data);
            }}
            pDic = _dic;
        }}
    }}
    #endregion
    """

conststr_commonRegion = """
    #region Common

    private static bool isInIt = false;

    {sheet_assign}
    
    public static void InIt(Dictionary<string,string[][]> strDic)
    {{
        if(isInIt)
            return;
        
        {sheet_init}

        isInIt = true;
    }}
    
    #endregion
"""

conststr_commonRegion_sheetInit = """
        string str{sheet_name} = "{sheet_name}";
        if (strDic.ContainsKey(str{sheet_name}))
        {{
            {sheet_name} = new {sheet_name}_Holder(strDic[str{sheet_name}]);
        }}
"""

conststr_commonRegion_sheetAssign = """
        public static {sheet_name}_Holder {sheet_name};
"""

def makeScriptByXlsx(wb_obj, table_name):
    str_generate = conststr_comment + conststr_namespace    
    str_scriptname = table_name + ".cs"
    allsheet = wb_obj.worksheets;


    str_allSheets = ""
    str_common = ""
    str_common_sheet_assign = ""
    str_common_sheet_getallvars = ""
    str_common_sheet_init = ""
    for sheet_obj in allsheet:
        if("Temp" in sheet_obj.title):
            continue

        str_sheet = ""
        str_sheet_data_struct = ""
        str_sheet_data_init = ""
        str_sheet_data_debug = ""

        #string_a += sheet_obj.title + '\t\t'
        m_column = sheet_obj.max_column
        for i in range(1, m_column + 1):
            if(sheet_obj.cell(row = 1 , column = i).value is None or sheet_obj.cell(row = 2 , column = i).value is None):
                continue  
            cell_type = str(sheet_obj.cell(row = 1 , column = i).value).upper()
            cell_varName = str(sheet_obj.cell(row = 2 , column = i).value);

            print ("cell_type is " + cell_type)
            print ("cell_valName is " + cell_varName)
            if cell_type == 'INT': 
                str_sheet_data_struct += '\t\t' + conststr_initInt.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoInt.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugInt.format(name = cell_varName)
            elif cell_type == "LONG":
                str_sheet_data_struct += '\t\t' + conststr_initLong.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoLong.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugLong.format(name = cell_varName)
            elif cell_type == "FLOAT":
                str_sheet_data_struct += '\t\t' + conststr_initFloat.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoFloat.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugFloat.format(name = cell_varName)
            elif cell_type == "DOUBLE":
                str_sheet_data_struct += '\t\t' + conststr_initDouble.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoDouble.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugDouble.format(name = cell_varName)
            elif cell_type == "STRING":
                str_sheet_data_struct += '\t\t' + conststr_initString.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoString.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugString.format(name = cell_varName)
            elif cell_type == "BOOL":
                str_sheet_data_struct += '\t\t' + conststr_initBool.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoBool.format(cell_varName,   i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugBool.format(name = cell_varName)
            elif cell_type == "XINT":
                str_sheet_data_struct += '\t\t' + conststr_initXInt.format(name = cell_varName)
                str_sheet_data_init   += '\t\t\t\t' + conststr_data_cvtoXint.format(cell_varName, i-1)
                str_sheet_data_debug += '\t\t' + conststr_debugXint.format(name = cell_varName)
            else:
                print("Custom Warn : type is anavaliable")
        str_sheet = conststr_sheetRegion.format(sheet_name = sheet_obj.title, data_struct = str_sheet_data_struct, data_init = str_sheet_data_init, data_debug = str_sheet_data_debug)
        str_allSheets += str_sheet;
        str_common_sheet_assign += conststr_commonRegion_sheetAssign.format(sheet_name = sheet_obj.title)
        str_common_sheet_init += conststr_commonRegion_sheetInit.format(sheet_name = sheet_obj.title)

    str_common = conststr_commonRegion.format(sheet_assign = str_common_sheet_assign, sheet_init = str_common_sheet_init)
    str_generate += conststr_tableClass.format(table_name = table_name, content = str_allSheets + str_common)
    ##
    file = open(PathManage.script_path + "/" + str_scriptname, "w+")
    file.write(str_generate)
    # file_proeject = open(PathManage.project_script_path + "/" + str_scriptname, "w+")
    # file_proeject.write(str_generate)